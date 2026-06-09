from __future__ import annotations

from io import BytesIO
from typing import BinaryIO

import pandas as pd
from openpyxl import load_workbook

SUMMARY_SHEET_NAME = "Podsumowanie"


def _file_to_bytes(file: BinaryIO | bytes) -> bytes:
    if isinstance(file, bytes):
        return file
    if hasattr(file, "getvalue"):
        return file.getvalue()
    current_position = file.tell() if hasattr(file, "tell") else None
    if hasattr(file, "seek"):
        file.seek(0)
    data = file.read()
    if current_position is not None and hasattr(file, "seek"):
        file.seek(current_position)
    return data


def load_workbook_from_upload(uploaded_file: BinaryIO):
    """Load an XLSX workbook from a Streamlit upload object."""
    return load_workbook(BytesIO(_file_to_bytes(uploaded_file)))


def get_product_sheets(workbook) -> list[str]:
    """Return product batch sheets, ignoring the summary sheet."""
    return [sheet for sheet in workbook.sheetnames if sheet.strip().lower() != SUMMARY_SHEET_NAME.lower()]


def read_sheet_to_dataframe(uploaded_file: BinaryIO | bytes, sheet_name: str) -> pd.DataFrame:
    """Read a selected sheet as a DataFrame and normalize missing values."""
    df = pd.read_excel(BytesIO(_file_to_bytes(uploaded_file)), sheet_name=sheet_name, dtype={"id_product": object})
    return df.fillna("")


def update_product_description(
    df: pd.DataFrame,
    id_product: str | int,
    description_short: str,
    description: str,
) -> pd.DataFrame:
    """Update description fields for one product matched by id_product."""
    updated = df.copy().fillna("")

    for column in ("description", "description_short"):
        if column not in updated.columns:
            updated[column] = ""

    if "id_product" not in updated.columns:
        raise ValueError("Arkusz nie zawiera wymaganej kolumny id_product.")

    id_as_text = str(id_product).strip()
    ids = updated["id_product"].astype(str).str.strip()
    mask = ids == id_as_text

    if not mask.any():
        raise ValueError(f"Nie znaleziono produktu o id_product={id_product}.")

    updated.loc[mask, "description_short"] = description_short
    updated.loc[mask, "description"] = description
    return updated


def write_updated_excel(original_file: BinaryIO | bytes, updated_sheets: dict[str, pd.DataFrame]) -> bytes:
    """Write all original sheets to XLSX, replacing only sheets present in updated_sheets."""
    original_bytes = _file_to_bytes(original_file)
    excel_file = pd.ExcelFile(BytesIO(original_bytes))
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name in excel_file.sheet_names:
            if sheet_name in updated_sheets:
                df = updated_sheets[sheet_name].fillna("")
            else:
                df = pd.read_excel(BytesIO(original_bytes), sheet_name=sheet_name).fillna("")
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    return output.getvalue()
